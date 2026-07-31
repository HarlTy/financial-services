# Build the retirement strategy workbook.
#
# Successor to the local-only build_v2.py, with one difference: this file holds
# ZERO legal literals and ZERO personal literals. Every limit, threshold, rate,
# and citation comes from parameters.json through paramlayer; every balance,
# birth year, and model assumption comes from an inputs JSON file. Structure,
# formulas, and formatting are otherwise unchanged, and test_identity.py proves
# it cell by cell.
#
# The prose is parameterised too, not just the cells. A note quoting an IRMAA
# threshold is as much a legal literal as the formula it sits beside, and a
# section header reciting opening balances is personal data in a public file.
# Both are composed at build time.
#
# What deliberately stays here: presentation (fonts, fills, number formats),
# layout anchors, flag strings, arithmetic tolerances, research references
# ("Bengen 4%" names a published rule of thumb, not this model's rate), and
# citations that name no value and parameterise nothing.
#
# Conventions (xlsx skill): blue = editable input, black = formula, green = pure
# cross-sheet link, yellow fill = key inputs. Growth accrues on beginning-of-year
# balances; contributions/draws/conversions post at year-end (original model's
# convention, preserved for comparability).
import argparse
import datetime as dt
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
if str(HERE) not in sys.path:
    sys.path.insert(0, str(HERE))
import paramlayer  # noqa: E402

DEFAULT_OUT = HERE.parents[1] / "retirement-strategy.xlsx"
DEFAULT_INPUTS = HERE / "inputs.local.json"
EXAMPLE_INPUTS = HERE / "inputs.example.json"

PREPARED = "2026-07-28"  # document date, not a legal or personal fact

ARIAL = "Arial"
F_TITLE = Font(name=ARIAL, size=12, bold=True)
F_NOTE = Font(name=ARIAL, size=9, italic=True, color="595959")
F_HDR = Font(name=ARIAL, size=10, bold=True)
F_IN = Font(name=ARIAL, size=10, color="0000FF")      # blue input
F_FX = Font(name=ARIAL, size=10, color="000000")      # black formula
F_LN = Font(name=ARIAL, size=10, color="008000")      # green cross-sheet link
F_SEC = Font(name=ARIAL, size=10, bold=True, color="1F3864")

FILL_HDR = PatternFill("solid", fgColor="D9D9D9")
FILL_KEY = PatternFill("solid", fgColor="FFFF00")

MONEY = '"$"#,##0'
MONEY2 = '"$"#,##0.00'
PCT = '0.0%'
PCT2 = '0.00%'
YEAR = '0'
NUM1 = '0.0'

WRAP = Alignment(wrap_text=True, vertical="top")


# ---------------------------------------------------------------------------
# Rendering helpers. These reproduce the exact shapes the workbook's prose uses.
# ---------------------------------------------------------------------------

def usd0(v):
    return f"${v:,.0f}"


def num(v):
    """Thousands-separated, decimals only when they carry meaning: a whole
    number loses its '.00', a fractional one keeps its digits."""
    s = f"{v:,.2f}"
    return s.rstrip("0").rstrip(".") if "." in s else s


def pct(v):
    """'6.5%' / '3%' / '9.9%' -- %g absorbs binary-float noise like 0.2*100."""
    return f"{v * 100:g}%"


def usd_k(v):
    return f"${v / 1000:g}k"


def usd_m(v):
    return f"${v / 1_000_000:g}M"


_WORDS = {1: "one", 2: "two", 3: "three", 4: "four"}


def word(n):
    return _WORDS[int(n)]


def issued_display(iso):
    d = dt.date.fromisoformat(iso)
    return f"{d:%b} {d.day}, {d.year}"


def put(ws, ref, value, font=F_FX, fmt=None, fill=None, align=None):
    c = ws[ref]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    return c


def load_inputs(path):
    data = json.loads(Path(path).read_text(encoding="utf-8"))
    missing = [k for k in ("person", "household", "timeline", "balances",
                           "contributions", "assumptions", "swr", "gap")
               if k not in data]
    if missing:
        raise SystemExit(f"{path}: missing input sections: {', '.join(missing)}")
    return data


def build(out_path, inputs_path, parameters_path=None):
    inp = load_inputs(inputs_path)
    reg = paramlayer.load(Path(parameters_path) if parameters_path else None)

    # -- plane 2 ------------------------------------------------------------
    BIRTH = inp["person"]["birth_year"]
    FILING = inp["household"]["filing_status"]
    START_Y = inp["timeline"]["model_start_year"]
    RET_AGE = inp["timeline"]["retirement_age"]
    DRAW_Y = inp["timeline"]["draw_start_year"]
    END_Y = inp["timeline"]["model_end_year"]
    OB_TAXABLE = inp["balances"]["taxable"]
    OB_TAXDEF = inp["balances"]["tax_deferred"]
    OB_ROTHV = inp["balances"]["roth"]
    EMP_V = inp["contributions"]["employee_deferral_start"]
    ER_V = inp["contributions"]["employer_contribution_start"]
    DESTINATION = inp["contributions"]["destination"]
    RET_V = inp["assumptions"]["nominal_return"]
    INFL_V = inp["assumptions"]["inflation"]
    ESC_V = inp["assumptions"]["contribution_escalation"]
    IDX_V = inp["assumptions"]["limit_indexation"]
    IRMIDX_V = inp["assumptions"]["irmaa_indexation"]
    WAIDX_V = inp["assumptions"]["wa_deduction_indexation"]
    GAINF_V = inp["assumptions"]["gain_fraction"]
    SWR_METHOD = inp["swr"]["method"]
    SWR_V = inp["swr"]["initial_rate"]
    BAND_V = inp["swr"]["guardrail_band"]
    ADJ_V = inp["swr"]["guardrail_adjustment"]
    LEGACY_V = inp["swr"]["legacy_flat_draw"]
    SCEN_V = inp["gap"]["scenario"]
    CONV_V = inp["gap"]["roth_conversion_per_year"]
    CTAX_V = inp["gap"]["conversion_tax_rate"]
    OMAGI_V = inp["gap"]["other_magi_income"]

    LAST_CONTRIB = BIRTH + RET_AGE
    GAP_FIRST, GAP_LAST = LAST_CONTRIB + 1, DRAW_Y - 1
    END_AGE = END_Y - BIRTH

    AS_OF = dt.date(START_Y, 1, 1)

    # -- plane 1 ------------------------------------------------------------
    def val(pid):
        return reg.value_as_of(pid, AS_OF)

    # Parameters the model projects past are unreachable by as_of() -- that is
    # what their closed window means -- so their provenance comes through
    # latest_published_record instead.
    PROJECTED_PAST = {"us-wa.rcw82_87.standard_deduction",
                      "us-wa.rcw83_100.estate_exclusion_base"}

    def _rec(pid):
        if pid in PROJECTED_PAST:
            return reg.latest_published_record(pid)
        return reg.record_for(pid, AS_OF)

    def pro(pid, field):
        return _rec(pid)["provenance"][field]

    def conf(pid):
        return pro(pid, "confidence").upper()

    def tyear(pid):
        return reg.as_of(pid, AS_OF)["tax_year"]

    def const(pid, name):
        return reg.as_of(pid, AS_OF)["constants"][name]

    def sec_of(pid):
        """'IRC 402(g)' -> '402(g)'."""
        return pro(pid, "code_cite").split(" ", 1)[1]

    ID_402G = "us.irc.402g.elective_deferral"
    ID_CU50 = "us.irc.414v.catchup_50plus"
    ID_CU60 = "us.irc.414v.catchup_60_63"
    ID_415C = "us.irc.415c.annual_additions"
    ID_CUAGE = "us.irc.414v.catchup_age_rule"
    ID_ROTHCU = "us.secure2.603.roth_catchup_wage_threshold"
    ID_IRMAA_MFJ = "us.medicare.irmaa.tier1_threshold.mfj"
    ID_IRMAA_SGL = "us.medicare.irmaa.tier1_threshold.single"
    ID_LOOKBACK = "us.medicare.irmaa.lookback_years"
    ID_FREEZE = "us.medicare.irmaa.fifth_tier_freeze_year"
    ID_WARATE = "us-wa.rcw82_87.excise_rate_base"
    ID_WASUR = "us-wa.rcw82_87.excise_rate_surcharge"
    ID_WATHR = "us-wa.rcw82_87.surcharge_threshold"
    ID_WADED = "us-wa.rcw82_87.standard_deduction"
    ID_WAEST = "us-wa.rcw83_100.estate_exclusion_base"
    ID_RMD = "us.irc.401a9.rmd_applicable_age"
    ID_QCD = "us.irc.408d8.qcd_annual_limit"
    ID_QCDAGE = "us.irc.408d8.qcd_eligibility_age"
    ID_NIIT = "us.irc.1411.niit_threshold.mfj"
    ID_FRA = "us.ssa.full_retirement_age"
    ID_CLAIM_LO = "us.ssa.claim_age.earliest"
    ID_CLAIM_HI = "us.ssa.claim_age.delayed_credit_end"

    V402G, VCU50, VCU60, V415C = val(ID_402G), val(ID_CU50), val(ID_CU60), val(ID_415C)
    VROTHCU = val(ID_ROTHCU)
    VIRMAA_MFJ, VIRMAA_SGL = val(ID_IRMAA_MFJ), val(ID_IRMAA_SGL)
    VLOOKBACK = val(ID_LOOKBACK)
    VWARATE, VWASUR, VWATHR = val(ID_WARATE), val(ID_WASUR), val(ID_WATHR)
    VQCD, VQCDAGE, VNIIT = val(ID_QCD), val(ID_QCDAGE), val(ID_NIIT)
    VCLAIM_LO, VCLAIM_HI = val(ID_CLAIM_LO), val(ID_CLAIM_HI)

    AGE50 = const(ID_CUAGE, "catchup_50_from_age")
    AGE60 = const(ID_CUAGE, "enhanced_from_age")
    AGE63 = const(ID_CUAGE, "enhanced_to_age")
    RMD_COHORT = const(ID_RMD, "cohort_birth_year")
    RMD_AGE_HI = const(ID_RMD, "applicable_age_at_or_after")
    RMD_AGE_LO = const(ID_RMD, "applicable_age_before")
    RMD_PRIOR_FIRST = const(ID_RMD, "prior_cohort_first_birth_year")
    FRA_AGE = const(ID_FRA, "full_retirement_age")
    FRA_COHORT = const(ID_FRA, "cohort_birth_year")

    # The WA standard deduction and estate exclusion are the two parameters the
    # model must project PAST: the last published amount is indexed forward by a
    # plane-2 assumption. latest_published hands back the validity end alongside
    # the value, so the extrapolation is explicit rather than accidental.
    wa_ded_payload, wa_ded_until = reg.latest_published(ID_WADED)
    VWADED = wa_ded_payload["value"]
    WADED_TY = wa_ded_payload["tax_year"]
    assert wa_ded_until is not None, "a projected-past parameter must have a validity end"
    wa_est_payload, _ = reg.latest_published(ID_WAEST)
    VWAEST = wa_est_payload["value"]
    # The TY after the last published deduction is what the note tells a reader
    # to verify -- taken from the 'expected' record, not computed.
    WADED_EXPECT_TY = next(
        r["payload"]["tax_year"] for r in reg.records_for(ID_WADED)
        if r["status"] == "expected")

    ULT = {}
    for r in reg.records:
        pid = r["parameter_id"]
        if pid.startswith("us.treasreg.1_401a9_9.ult_divisor.age_"):
            ULT[int(pid.rsplit("_", 1)[1])] = r["payload"]["value"]
    ULT = dict(sorted(ULT.items()))
    ULT_ID0 = f"us.treasreg.1_401a9_9.ult_divisor.age_{min(ULT)}"
    ULT_FROM_Y = dt.date.fromisoformat(
        reg.record_for(ULT_ID0, AS_OF)["valid_from"]).year

    SOURCES = "; ".join([
        pro(ID_402G, "short_cite"),
        f"{pro(ID_IRMAA_MFJ, 'short_cite')} ({issued_display(pro(ID_IRMAA_MFJ, 'issued_on'))})",
        pro(ID_WARATE, "short_cite"),
        pro(ID_RMD, "short_cite"),
    ])

    # =====================================================================
    wb = openpyxl.Workbook()
    wsA = wb.active
    wsA.title = "Assumptions"
    wsAcc = wb.create_sheet("Accumulation")
    wsGap = wb.create_sheet("Gap")
    wsDd = wb.create_sheet("Drawdown")
    wsChk = wb.create_sheet("Checks")
    wsAdv = wb.create_sheet("Advisor")

    # =====================================================================
    # ASSUMPTIONS
    # =====================================================================
    put(wsA, "A1", "Retirement Strategy v2 - Assumptions", F_TITLE)
    put(wsA, "A2", "Legend: BLUE font = editable input | BLACK = formula | GREEN = link to another sheet | YELLOW fill = key inputs to revisit. Edit blue cells only.", F_NOTE)
    put(wsA, "A3", f"All dollars are nominal unless a column is labeled 'real ({START_Y} $)'. Every limit and threshold carries its tax year. Sources: {SOURCES}. VERIFY = confirm against primary source before it drives a decision.", F_NOTE)

    def label(row, text, note=""):
        put(wsA, f"A{row}", text, F_FX)
        if note:
            put(wsA, f"C{row}", note, F_NOTE, align=WRAP)

    def section(row, text):
        put(wsA, f"A{row}", text, F_SEC)

    section(5, "TIMELINE")
    label(6, "Model start year", f"From original model (first projection year {START_Y}).")
    put(wsA, "B6", START_Y, F_IN, YEAR)
    label(7, "Birth year", f"Derived from original: age {START_Y - BIRTH} in {START_Y}. Confirm.")
    put(wsA, "B7", BIRTH, F_IN, YEAR)
    label(8, "Age at model start")
    put(wsA, "B8", "=B6-B7", F_FX, YEAR)
    label(9, "Retirement age (last contribution year age)", f"Original contributions end in {LAST_CONTRIB} (age {RET_AGE}).")
    put(wsA, "B9", RET_AGE, F_IN, YEAR, FILL_KEY)
    label(10, "Last contribution year")
    put(wsA, "B10", "=B7+B9", F_FX, YEAR)
    label(11, "Draw start year", f"Original draws begin {DRAW_Y} (age {DRAW_Y - BIRTH}). Structural: Drawdown tab covers {DRAW_Y}-{END_Y}; set this within that range.")
    put(wsA, "B11", DRAW_Y, F_IN, YEAR, FILL_KEY)
    label(12, "Model end year", f"Structural - table lengths are fixed to {END_Y} (age {END_AGE}).")
    put(wsA, "B12", END_Y, F_FX, YEAR)

    section(14, "RETURNS AND INFLATION")
    label(15, "Nominal annual return", f"From original model (C2 = {pct(RET_V)}). Planning signal, not a forecast.")
    put(wsA, "B15", RET_V, F_IN, PCT2, FILL_KEY)
    label(16, "Inflation assumption", f"Assumption. Matches the {pct(ESC_V)} escalator the original intended but never applied (defect 1).")
    put(wsA, "B16", INFL_V, F_IN, PCT2, FILL_KEY)
    label(17, "Approx. real return (derived)")
    put(wsA, "B17", "=(1+B15)/(1+B16)-1", F_FX, PCT2)

    section(19, "CONTRIBUTIONS (fix for defect 1 - dead escalator)")
    label(20, f"Employee deferral, {START_Y}", f"TY{START_Y}. Split of the original combined {usd0(EMP_V + ER_V)} is ASSUMED - edit both cells to your actual split.")
    put(wsA, "B20", EMP_V, F_IN, MONEY)
    label(21, f"Employer contribution, {START_Y}", f"TY{START_Y}. See note above.")
    put(wsA, "B21", ER_V, F_IN, MONEY)
    label(22, f"Combined {START_Y} contribution (check)", f"Equals the original model's {usd0(EMP_V + ER_V)} 'Employer + Employee'.")
    put(wsA, "B22", "=B20+B21", F_FX, MONEY)
    label(23, "Contribution escalation rate", "The live escalator. Original had =C20*0.03+C20 anchored to a zero cell, so it never escalated.")
    put(wsA, "B23", ESC_V, F_IN, PCT2, FILL_KEY)
    label(24, "Contribution destination", "All contributions modeled as pre-tax into the tax-deferred bucket (matches original's single-account framing). Discuss Roth deferrals with advisor.")
    put(wsA, "B24", DESTINATION, F_IN)

    section(26, f"TY{tyear(ID_402G)} CONTRIBUTION LIMITS ({pro(ID_402G, 'short_cite')} - {conf(ID_402G)})")
    label(27, f"{sec_of(ID_402G)} elective deferral limit", f"TY{tyear(ID_402G)} {conf(ID_402G)}. Shared across ALL employers. {pro(ID_402G, 'code_cite')}.")
    put(wsA, "B27", V402G, F_IN, MONEY)
    label(28, f"Catch-up, age {AGE50}+", f"TY{tyear(ID_CU50)} {conf(ID_CU50)}. {pro(ID_CU50, 'code_cite')}.")
    put(wsA, "B28", VCU50, F_IN, MONEY)
    label(29, f"Enhanced catch-up, ages {AGE60}-{AGE63}", f"TY{tyear(ID_CU60)} {conf(ID_CU60)}. Replaces the {AGE50}+ catch-up in those years.")
    put(wsA, "B29", VCU60, F_IN, MONEY)
    label(30, f"{sec_of(ID_415C)} annual additions limit", f"TY{tyear(ID_415C)} {conf(ID_415C)}. Employee + employer, per unrelated employer.")
    put(wsA, "B30", V415C, F_IN, MONEY)
    label(31, "Limit indexation assumption", "Assumption. Real limits index in steps; modeled as smooth growth.")
    put(wsA, "B31", IDX_V, F_IN, PCT2)
    label(32, "Roth catch-up mandate note", f"From {tyear(ID_ROTHCU)}, if prior-year FICA wages from an employer exceeded {usd0(VROTHCU)}, catch-up must be Roth ({pro(ID_ROTHCU, 'short_cite')}). Not modeled - flag for advisor.")
    put(wsA, "B32", "See Advisor tab", F_FX)

    section(34, f"OPENING BALANCES, 1/1/{START_Y} (from original D5 = {num(OB_TAXDEF)} + {num(OB_TAXABLE)} + {num(OB_ROTHV)})")
    label(35, "Taxable brokerage", "Bucket mapping of the three components is ASSUMED - relabel to actuals.")
    put(wsA, "B35", OB_TAXABLE, F_IN, MONEY2)
    label(36, "Tax-deferred (401k/IRA)", "Largest component assumed to be the 401(k).")
    put(wsA, "B36", OB_TAXDEF, F_IN, MONEY2)
    label(37, "Roth", "")
    put(wsA, "B37", OB_ROTHV, F_IN, MONEY2)
    label(38, f"Total (check vs original ${num(OB_TAXABLE + OB_TAXDEF + OB_ROTHV)})")
    put(wsA, "B38", "=B35+B36+B37", F_FX, MONEY2)

    section(40, "DRAWDOWN METHOD (fixes defects 2 and 3)")
    label(41, "SWR method", "Fixed = Bengen-style: initial rate x balance at draw start, then inflation-adjusted. Guardrails = Guyton-Klinger-style bands (simplified).")
    put(wsA, "B41", SWR_METHOD, F_IN, None, FILL_KEY)
    label(42, "Initial withdrawal rate", f"A RATE, not a dollar amount (original's 'Withdrawal Rate' label held {usd0(LEGACY_V)} - defect 3). Bengen 4%; Morningstar 2025 base case 3.9%.")
    put(wsA, "B42", SWR_V, F_IN, PCT2, FILL_KEY)
    label(43, "Guardrail band (+/-)", "Guardrails only: allowed drift of current withdrawal rate around the initial rate.")
    put(wsA, "B43", BAND_V, F_IN, PCT)
    label(44, "Guardrail adjustment", "Guardrails only: cut/raise applied when a band is breached.")
    put(wsA, "B44", ADJ_V, F_IN, PCT)
    label(45, "Legacy flat draw (reference only)", f"The original model's {usd_k(LEGACY_V)} nominal draw. NOT used in any calculation; kept to compare against defect 2.")
    put(wsA, "B45", LEGACY_V, F_IN, MONEY)

    section(47, f"GAP YEARS {GAP_FIRST}-{GAP_LAST} (fix for defect 4)")
    label(48, "Gap scenario feeding Drawdown", "A = growth only, B = early draws, C = Roth conversion window.")
    put(wsA, "B48", SCEN_V, F_IN, None, FILL_KEY)
    label(49, "Roth conversion per gap year (Scenario C)", "Size with your CPA against current-year bracket tables - deliberately NOT computed from brackets here.")
    put(wsA, "B49", CONV_V, F_IN, MONEY, FILL_KEY)
    label(50, "Effective tax rate on conversions", "Assumption. Tax paid from the taxable bucket (preserves conversion value).")
    put(wsA, "B50", CTAX_V, F_IN, PCT, FILL_KEY)
    label(51, "Other MAGI income in gap years", "Assumption - interest/dividends/other income beside conversions.")
    put(wsA, "B51", OMAGI_V, F_IN, MONEY)
    label(52, "Filing status", "Skill default MFJ - confirm; IRMAA threshold below depends on it.")
    put(wsA, "B52", FILING, F_IN)
    label(53, f"IRMAA first-tier MAGI threshold, {tyear(ID_IRMAA_MFJ)}", f"{tyear(ID_IRMAA_MFJ)} {conf(ID_IRMAA_MFJ)} ({pro(ID_IRMAA_MFJ, 'issuer')}): {usd0(VIRMAA_MFJ)} MFJ / {usd0(VIRMAA_SGL)} single. Assessed per person, {word(VLOOKBACK)}-year lookback, CLIFF - $1 over triggers the full surcharge.")
    put(wsA, "B53", f'=IF(B52="MFJ",{VIRMAA_MFJ:.0f},{VIRMAA_SGL:.0f})', F_FX, MONEY)
    label(54, "IRMAA threshold indexation assumption", f"Assumption. First four tiers are indexed (fifth frozen until {val(ID_FREEZE):.0f}).")
    put(wsA, "B54", IRMIDX_V, F_IN, PCT2)

    section(56, f"WASHINGTON STATE ({pro(ID_WARATE, 'short_cite')})")
    label(57, "WA capital-gains excise rate", f"{conf(ID_WARATE)}. {pct(VWARATE)} on WA long-term gains above the standard deduction. ({pct(VWARATE + VWASUR)} tier above {usd_m(VWATHR)} of gains not modeled - flag if a single-year realization approaches {usd_m(VWATHR)}.)")
    put(wsA, "B57", VWARATE, F_IN, PCT)
    label(58, f"WA CG standard deduction (TY{WADED_TY})", f"TY{WADED_TY} = {usd0(VWADED)}, CPI-indexed. {conf(ID_WADED)} the TY{WADED_EXPECT_TY} amount at {pro(ID_WADED, 'issuer')} before relying on a specific year.")
    put(wsA, "B58", VWADED, F_IN, MONEY)
    label(59, "WA deduction indexation assumption", "Assumption (statute indexes to CPI).")
    put(wsA, "B59", WAIDX_V, F_IN, PCT2)
    label(60, "Gain fraction of taxable withdrawals", "Assumption: share of each taxable-bucket withdrawal that is realized long-term gain.")
    put(wsA, "B60", GAINF_V, F_IN, PCT)
    label(61, "WA exemption note", "Retirement-account assets are EXEMPT from the WA excise - only the taxable bucket is checked.")
    put(wsA, "B61", "Taxable bucket only", F_FX)

    section(63, f"RMD ({pro(ID_RMD, 'short_cite')} / {pro(ID_RMD, 'code_cite')})")
    label(64, "RMD start age", f"{conf(ID_RMD)}: {RMD_AGE_LO:.0f} if born {RMD_PRIOR_FIRST:.0f}-{RMD_COHORT - 1:.0f}; {RMD_AGE_HI:.0f} if born {RMD_COHORT:.0f}+.")
    put(wsA, "B64", f"=IF(B7>={RMD_COHORT:.0f},{RMD_AGE_HI:.0f},{RMD_AGE_LO:.0f})", F_FX, YEAR)
    label(65, "First RMD year")
    put(wsA, "B65", "=B7+B64", F_FX, YEAR)
    label(66, "Uniform Lifetime Table (divisors)", f"IRS Uniform Lifetime Table, {pro(ULT_ID0, 'code_cite')} (post-{ULT_FROM_Y}). {conf(ULT_ID0)} against current IRS Pub. 590-B before relying.")
    put(wsA, "A67", "Age", F_HDR, None, FILL_HDR)
    put(wsA, "B67", "Divisor", F_HDR, None, FILL_HDR)
    r = 68
    for age, div in ULT.items():
        put(wsA, f"A{r}", age, F_IN, YEAR)
        put(wsA, f"B{r}", div, F_IN, NUM1)
        r += 1
    ULT_FIRST, ULT_LAST = 68, r - 1

    dv1 = DataValidation(type="list", formula1='"Fixed,Guardrails"', allow_blank=False)
    wsA.add_data_validation(dv1); dv1.add("B41")
    dv2 = DataValidation(type="list", formula1='"A,B,C"', allow_blank=False)
    wsA.add_data_validation(dv2); dv2.add("B48")
    dv3 = DataValidation(type="list", formula1='"MFJ,Single"', allow_blank=False)
    wsA.add_data_validation(dv3); dv3.add("B52")

    wsA.column_dimensions["A"].width = 44
    wsA.column_dimensions["B"].width = 16
    wsA.column_dimensions["C"].width = 90

    # Shorthand refs into Assumptions
    START = "Assumptions!$B$6"; BIRTH_R = "Assumptions!$B$7"; RET_Y = "Assumptions!$B$10"
    DRAW_YR = "Assumptions!$B$11"; RET_RATE = "Assumptions!$B$15"; INFL = "Assumptions!$B$16"
    EMP = "Assumptions!$B$20"; ER = "Assumptions!$B$21"; ESC = "Assumptions!$B$23"
    L402 = "Assumptions!$B$27"; CU50 = "Assumptions!$B$28"; CU60 = "Assumptions!$B$29"
    L415 = "Assumptions!$B$30"; IDX = "Assumptions!$B$31"
    OB_TAX = "Assumptions!$B$35"; OB_DEF = "Assumptions!$B$36"; OB_ROTH = "Assumptions!$B$37"
    METHOD = "Assumptions!$B$41"; SWR = "Assumptions!$B$42"; BAND = "Assumptions!$B$43"
    ADJ = "Assumptions!$B$44"
    SCEN = "Assumptions!$B$48"; CONV = "Assumptions!$B$49"; CTAX = "Assumptions!$B$50"
    OMAGI = "Assumptions!$B$51"; IRMAA = "Assumptions!$B$53"; IRMIDX = "Assumptions!$B$54"
    WARATE = "Assumptions!$B$57"; WADED = "Assumptions!$B$58"; WAIDX = "Assumptions!$B$59"
    GAINF = "Assumptions!$B$60"
    RMDAGE = "Assumptions!$B$64"
    ULT_A = f"Assumptions!$A${ULT_FIRST}:$A${ULT_LAST}"
    ULT_B = f"Assumptions!$B${ULT_FIRST}:$B${ULT_LAST}"

    # =====================================================================
    # ACCUMULATION
    # =====================================================================
    put(wsAcc, "A1", f"Accumulation {START_Y}-{LAST_CONTRIB} (ages {START_Y - BIRTH}-{RET_AGE})", F_TITLE)
    put(wsAcc, "A2", "Growth accrues on beginning-of-year balances; contributions post at year-end (original model's convention). Contributions stop after the last contribution year on Assumptions.", F_NOTE)
    put(wsAcc, "A3", f"Caps: employee deferral capped at indexed {sec_of(ID_402G)} + age-based catch-up; total capped at indexed {sec_of(ID_415C)}. 'CAPPED' appears in the flag column when a cap binds.", F_NOTE)
    acc_hdrs = ["Year", "Age", "Beg Taxable", "Beg Tax-Def", "Beg Roth", "Beg Total",
                "Deferral (pre-cap)",
                f"Deferral cap ({sec_of(ID_402G).replace('(', '').replace(')', '')}+CU, idx)",
                "Deferral (capped)",
                "Employer", f"{sec_of(ID_415C)} cap (idx)", "Total contrib (capped)", "Cap flag",
                "Growth Taxable", "Growth Tax-Def", "Growth Roth",
                "End Taxable", "End Tax-Def", "End Roth", "End Total (nominal)",
                f"End Total real ({START_Y} $)", f"Contrib real ({START_Y} $)"]
    for i, h in enumerate(acc_hdrs, 1):
        put(wsAcc, f"{get_column_letter(i)}4", h, F_HDR, None, FILL_HDR, WRAP)

    ACC_FIRST = 5
    ACC_LAST = ACC_FIRST + (LAST_CONTRIB - START_Y)
    for rr in range(ACC_FIRST, ACC_LAST + 1):
        if rr == ACC_FIRST:
            put(wsAcc, f"A{rr}", f"={START}", F_LN, YEAR)
            put(wsAcc, f"C{rr}", f"={OB_TAX}", F_LN, MONEY)
            put(wsAcc, f"D{rr}", f"={OB_DEF}", F_LN, MONEY)
            put(wsAcc, f"E{rr}", f"={OB_ROTH}", F_LN, MONEY)
        else:
            put(wsAcc, f"A{rr}", f"=A{rr-1}+1", F_FX, YEAR)
            put(wsAcc, f"C{rr}", f"=Q{rr-1}", F_FX, MONEY)
            put(wsAcc, f"D{rr}", f"=R{rr-1}", F_FX, MONEY)
            put(wsAcc, f"E{rr}", f"=S{rr-1}", F_FX, MONEY)
        put(wsAcc, f"B{rr}", f"=A{rr}-{BIRTH_R}", F_FX, YEAR)
        put(wsAcc, f"F{rr}", f"=C{rr}+D{rr}+E{rr}", F_FX, MONEY)
        put(wsAcc, f"G{rr}", f"=IF(A{rr}<={RET_Y},{EMP}*(1+{ESC})^(A{rr}-{START}),0)", F_FX, MONEY)
        put(wsAcc, f"H{rr}", f"=({L402}+IF(AND(B{rr}>={AGE60:.0f},B{rr}<={AGE63:.0f}),{CU60},IF(B{rr}>={AGE50:.0f},{CU50},0)))*(1+{IDX})^(A{rr}-{START})", F_FX, MONEY)
        put(wsAcc, f"I{rr}", f"=MIN(G{rr},H{rr})", F_FX, MONEY)
        put(wsAcc, f"J{rr}", f"=IF(A{rr}<={RET_Y},{ER}*(1+{ESC})^(A{rr}-{START}),0)", F_FX, MONEY)
        put(wsAcc, f"K{rr}", f"={L415}*(1+{IDX})^(A{rr}-{START})", F_FX, MONEY)
        put(wsAcc, f"L{rr}", f"=MIN(I{rr}+J{rr},K{rr})", F_FX, MONEY)
        put(wsAcc, f"M{rr}", f'=IF(L{rr}<G{rr}+J{rr}-0.005,"CAPPED","-")', F_FX)
        put(wsAcc, f"N{rr}", f"=C{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsAcc, f"O{rr}", f"=D{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsAcc, f"P{rr}", f"=E{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsAcc, f"Q{rr}", f"=C{rr}+N{rr}", F_FX, MONEY)
        put(wsAcc, f"R{rr}", f"=D{rr}+O{rr}+L{rr}", F_FX, MONEY)
        put(wsAcc, f"S{rr}", f"=E{rr}+P{rr}", F_FX, MONEY)
        put(wsAcc, f"T{rr}", f"=Q{rr}+R{rr}+S{rr}", F_FX, MONEY)
        put(wsAcc, f"U{rr}", f"=T{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)
        put(wsAcc, f"V{rr}", f"=L{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)

    wsAcc.freeze_panes = "C5"
    for col, w in {"A": 7, "B": 5}.items():
        wsAcc.column_dimensions[col].width = w
    for i in range(3, 23):
        wsAcc.column_dimensions[get_column_letter(i)].width = 13

    # =====================================================================
    # GAP
    # =====================================================================
    put(wsGap, "A1", f"Gap years {GAP_FIRST}-{GAP_LAST} (ages {GAP_FIRST - BIRTH}-{GAP_LAST - BIRTH}) - three explicit scenarios (fix for defect 4)", F_TITLE)
    put(wsGap, "A2", f"A = growth only. B = early draws (initial rate x {GAP_FIRST} balance, then inflation-adjusted; taxable -> tax-deferred -> Roth; no RMDs before the RMD start age). C = Roth conversion window with IRMAA {word(VLOOKBACK)}-year lookahead check.", F_NOTE)
    put(wsGap, "A3", "The scenario selected on Assumptions (B48) feeds the Drawdown tab via the handoff block at the bottom.", F_NOTE)

    put(wsGap, "A4", "SCENARIO A - GROWTH ONLY", F_SEC)
    ga_hdrs = ["Year", "Age", "Beg Taxable", "Beg Tax-Def", "Beg Roth", "Beg Total",
               "Growth Taxable", "Growth Tax-Def", "Growth Roth",
               "End Taxable", "End Tax-Def", "End Roth", "End Total",
               f"End Total real ({START_Y} $)"]
    for i, h in enumerate(ga_hdrs, 1):
        put(wsGap, f"{get_column_letter(i)}5", h, F_HDR, None, FILL_HDR, WRAP)
    GA_FIRST = 6
    GA_LAST = GA_FIRST + (GAP_LAST - GAP_FIRST)
    for rr in range(GA_FIRST, GA_LAST + 1):
        if rr == GA_FIRST:
            put(wsGap, f"A{rr}", f"=Accumulation!A{ACC_LAST}+1", F_LN, YEAR)
            put(wsGap, f"C{rr}", f"=Accumulation!Q{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"D{rr}", f"=Accumulation!R{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"E{rr}", f"=Accumulation!S{ACC_LAST}", F_LN, MONEY)
        else:
            put(wsGap, f"A{rr}", f"=A{rr-1}+1", F_FX, YEAR)
            put(wsGap, f"C{rr}", f"=J{rr-1}", F_FX, MONEY)
            put(wsGap, f"D{rr}", f"=K{rr-1}", F_FX, MONEY)
            put(wsGap, f"E{rr}", f"=L{rr-1}", F_FX, MONEY)
        put(wsGap, f"B{rr}", f"=A{rr}-{BIRTH_R}", F_FX, YEAR)
        put(wsGap, f"F{rr}", f"=C{rr}+D{rr}+E{rr}", F_FX, MONEY)
        put(wsGap, f"G{rr}", f"=C{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"H{rr}", f"=D{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"I{rr}", f"=E{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"J{rr}", f"=C{rr}+G{rr}", F_FX, MONEY)
        put(wsGap, f"K{rr}", f"=D{rr}+H{rr}", F_FX, MONEY)
        put(wsGap, f"L{rr}", f"=E{rr}+I{rr}", F_FX, MONEY)
        put(wsGap, f"M{rr}", f"=J{rr}+K{rr}+L{rr}", F_FX, MONEY)
        put(wsGap, f"N{rr}", f"=M{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)

    GB_SEC = GA_LAST + 2
    put(wsGap, f"A{GB_SEC}", f"SCENARIO B - EARLY DRAWS (draws begin {GAP_FIRST} instead of {DRAW_Y})", F_SEC)
    gb_hdrs = ["Year", "Age", "Beg Taxable", "Beg Tax-Def", "Beg Roth", "Beg Total",
               "Draw target", "From Taxable", "From Tax-Def", "From Roth", "Spending",
               "Growth Taxable", "Growth Tax-Def", "Growth Roth",
               "End Taxable", "End Tax-Def", "End Roth", "End Total",
               f"End Total real ({START_Y} $)"]
    GB_HDR_ROW = GB_SEC + 1
    for i, h in enumerate(gb_hdrs, 1):
        put(wsGap, f"{get_column_letter(i)}{GB_HDR_ROW}", h, F_HDR, None, FILL_HDR, WRAP)
    GB_FIRST = GB_HDR_ROW + 1
    GB_LAST = GB_FIRST + (GAP_LAST - GAP_FIRST)
    GB_OFFSET = GB_FIRST - GA_FIRST
    for rr in range(GB_FIRST, GB_LAST + 1):
        ra = rr - GB_OFFSET
        put(wsGap, f"A{rr}", f"=A{ra}", F_FX, YEAR)
        put(wsGap, f"B{rr}", f"=B{ra}", F_FX, YEAR)
        if rr == GB_FIRST:
            put(wsGap, f"C{rr}", f"=Accumulation!Q{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"D{rr}", f"=Accumulation!R{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"E{rr}", f"=Accumulation!S{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"G{rr}", f"={SWR}*F{rr}", F_FX, MONEY)
        else:
            put(wsGap, f"C{rr}", f"=O{rr-1}", F_FX, MONEY)
            put(wsGap, f"D{rr}", f"=P{rr-1}", F_FX, MONEY)
            put(wsGap, f"E{rr}", f"=Q{rr-1}", F_FX, MONEY)
            put(wsGap, f"G{rr}", f"=G{rr-1}*(1+{INFL})", F_FX, MONEY)
        put(wsGap, f"F{rr}", f"=C{rr}+D{rr}+E{rr}", F_FX, MONEY)
        put(wsGap, f"H{rr}", f"=MIN(G{rr},C{rr})", F_FX, MONEY)
        put(wsGap, f"I{rr}", f"=MIN(G{rr}-H{rr},D{rr})", F_FX, MONEY)
        put(wsGap, f"J{rr}", f"=MIN(G{rr}-H{rr}-I{rr},E{rr})", F_FX, MONEY)
        put(wsGap, f"K{rr}", f"=H{rr}+I{rr}+J{rr}", F_FX, MONEY)
        put(wsGap, f"L{rr}", f"=C{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"M{rr}", f"=D{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"N{rr}", f"=E{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"O{rr}", f"=C{rr}+L{rr}-H{rr}", F_FX, MONEY)
        put(wsGap, f"P{rr}", f"=D{rr}+M{rr}-I{rr}", F_FX, MONEY)
        put(wsGap, f"Q{rr}", f"=E{rr}+N{rr}-J{rr}", F_FX, MONEY)
        put(wsGap, f"R{rr}", f"=O{rr}+P{rr}+Q{rr}", F_FX, MONEY)
        put(wsGap, f"S{rr}", f"=R{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)

    GC_SEC = GB_LAST + 2
    put(wsGap, f"A{GC_SEC}", f"SCENARIO C - ROTH CONVERSION WINDOW (with IRMAA {word(VLOOKBACK)}-year lookahead check)", F_SEC)
    gc_hdrs = ["Year", "Age", "Beg Taxable", "Beg Tax-Def", "Beg Roth", "Beg Total",
               "Conversion", "Tax on conversion (from taxable)",
               "Growth Taxable", "Growth Tax-Def", "Growth Roth",
               "End Taxable", "End Tax-Def", "End Roth", "End Total",
               f"End Total real ({START_Y} $)",
               "MAGI proxy", "IRMAA tier-1 threshold (idx)",
               f"IRMAA check (surcharge lands Year+{VLOOKBACK:.0f})"]
    GC_HDR_ROW = GC_SEC + 1
    for i, h in enumerate(gc_hdrs, 1):
        put(wsGap, f"{get_column_letter(i)}{GC_HDR_ROW}", h, F_HDR, None, FILL_HDR, WRAP)
    GC_FIRST = GC_HDR_ROW + 1
    GC_LAST = GC_FIRST + (GAP_LAST - GAP_FIRST)
    GC_OFFSET = GC_FIRST - GA_FIRST
    for rr in range(GC_FIRST, GC_LAST + 1):
        ra = rr - GC_OFFSET
        put(wsGap, f"A{rr}", f"=A{ra}", F_FX, YEAR)
        put(wsGap, f"B{rr}", f"=B{ra}", F_FX, YEAR)
        if rr == GC_FIRST:
            put(wsGap, f"C{rr}", f"=Accumulation!Q{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"D{rr}", f"=Accumulation!R{ACC_LAST}", F_LN, MONEY)
            put(wsGap, f"E{rr}", f"=Accumulation!S{ACC_LAST}", F_LN, MONEY)
        else:
            put(wsGap, f"C{rr}", f"=L{rr-1}", F_FX, MONEY)
            put(wsGap, f"D{rr}", f"=M{rr-1}", F_FX, MONEY)
            put(wsGap, f"E{rr}", f"=N{rr-1}", F_FX, MONEY)
        put(wsGap, f"F{rr}", f"=C{rr}+D{rr}+E{rr}", F_FX, MONEY)
        put(wsGap, f"G{rr}", f"=MIN({CONV},D{rr})", F_FX, MONEY)
        put(wsGap, f"H{rr}", f"=MIN(G{rr}*{CTAX},C{rr}+I{rr})", F_FX, MONEY)
        put(wsGap, f"I{rr}", f"=C{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"J{rr}", f"=D{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"K{rr}", f"=E{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsGap, f"L{rr}", f"=C{rr}+I{rr}-H{rr}", F_FX, MONEY)
        put(wsGap, f"M{rr}", f"=D{rr}+J{rr}-G{rr}", F_FX, MONEY)
        put(wsGap, f"N{rr}", f"=E{rr}+K{rr}+G{rr}", F_FX, MONEY)
        put(wsGap, f"O{rr}", f"=L{rr}+M{rr}+N{rr}", F_FX, MONEY)
        put(wsGap, f"P{rr}", f"=O{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)
        put(wsGap, f"Q{rr}", f"=G{rr}+{OMAGI}", F_FX, MONEY)
        put(wsGap, f"R{rr}", f"={IRMAA}*(1+{IRMIDX})^(A{rr}-{START})", F_FX, MONEY)
        put(wsGap, f"S{rr}", f'=IF(Q{rr}>R{rr},"YES - IRMAA surcharge in "&(A{rr}+{VLOOKBACK:.0f}),"no")', F_FX)

    HANDOFF = GC_LAST + 2
    put(wsGap, f"A{HANDOFF}", f"HANDOFF TO DRAWDOWN - end-{GAP_LAST} balances of the scenario selected on Assumptions (B48)", F_SEC)
    put(wsGap, f"A{HANDOFF+1}", "Selected:", F_FX)
    put(wsGap, f"B{HANDOFF+1}", f"={SCEN}", F_LN)
    put(wsGap, f"A{HANDOFF+2}", "Feed:", F_FX)
    put(wsGap, f"B{HANDOFF+2}", "Taxable", F_HDR)
    put(wsGap, f"C{HANDOFF+2}", "Tax-deferred", F_HDR)
    put(wsGap, f"D{HANDOFF+2}", "Roth", F_HDR)
    FEED = HANDOFF + 3
    put(wsGap, f"B{FEED}", f'=IF({SCEN}="A",J{GA_LAST},IF({SCEN}="B",O{GB_LAST},L{GC_LAST}))', F_FX, MONEY)
    put(wsGap, f"C{FEED}", f'=IF({SCEN}="A",K{GA_LAST},IF({SCEN}="B",P{GB_LAST},M{GC_LAST}))', F_FX, MONEY)
    put(wsGap, f"D{FEED}", f'=IF({SCEN}="A",L{GA_LAST},IF({SCEN}="B",Q{GB_LAST},N{GC_LAST}))', F_FX, MONEY)

    wsGap.column_dimensions["A"].width = 9
    wsGap.column_dimensions["B"].width = 12
    for i in range(3, 20):
        wsGap.column_dimensions[get_column_letter(i)].width = 13
    wsGap.column_dimensions["S"].width = 26

    # =====================================================================
    # DRAWDOWN
    # =====================================================================
    put(wsDd, "A1", f"Drawdown {DRAW_Y}-{END_Y} (ages {DRAW_Y - BIRTH}-{END_AGE}) - inflation-escalating draws (fix for defect 2)", F_TITLE)
    put(wsDd, "A2", "Draw = initial rate x balance at draw start, then Fixed (inflation-adjusted, Bengen) or Guardrails (Guyton-Klinger-style, simplified: proposed draw's rate vs band around the initial rate; breach cuts/raises by the adjustment). Draws are GROSS of income tax.", F_NOTE)
    put(wsDd, "A3", "Sequencing: RMD from tax-deferred first (mandatory), then taxable, then tax-deferred, then Roth. Excess RMD above the draw target is reinvested into the taxable bucket. WA excise check applies to taxable-bucket realizations only.", F_NOTE)
    dd_hdrs = ["Year", "Age", "Beg Taxable", "Beg Tax-Def", "Beg Roth", "Beg Total",
               "Draw target", "RMD required", "RMD taken", "From Taxable", "From Tax-Def (extra)",
               "From Roth", "Spending funded", "Excess RMD to Taxable",
               "Growth Taxable", "Growth Tax-Def", "Growth Roth",
               "End Taxable", "End Tax-Def", "End Roth", "End Total (nominal)",
               f"End Total real ({START_Y} $)", f"Draw real ({START_Y} $)",
               "LTCG realized (taxable)", "WA std deduction (idx)", "WA excise est.", "WA flag"]
    for i, h in enumerate(dd_hdrs, 1):
        put(wsDd, f"{get_column_letter(i)}4", h, F_HDR, None, FILL_HDR, WRAP)

    DD_FIRST = 5
    DD_LAST = DD_FIRST + (END_Y - DRAW_Y)
    for rr in range(DD_FIRST, DD_LAST + 1):
        if rr == DD_FIRST:
            put(wsDd, f"A{rr}", f"=Gap!A{GA_LAST}+1", F_LN, YEAR)
            put(wsDd, f"C{rr}", f"=Gap!B{FEED}", F_LN, MONEY)
            put(wsDd, f"D{rr}", f"=Gap!C{FEED}", F_LN, MONEY)
            put(wsDd, f"E{rr}", f"=Gap!D{FEED}", F_LN, MONEY)
            put(wsDd, f"G{rr}", f"=IF(A{rr}<{DRAW_YR},0,IF(A{rr}={DRAW_YR},{SWR}*F{rr},0))", F_FX, MONEY)
        else:
            put(wsDd, f"A{rr}", f"=A{rr-1}+1", F_FX, YEAR)
            put(wsDd, f"C{rr}", f"=R{rr-1}", F_FX, MONEY)
            put(wsDd, f"D{rr}", f"=S{rr-1}", F_FX, MONEY)
            put(wsDd, f"E{rr}", f"=T{rr-1}", F_FX, MONEY)
            g_fixed = f"G{rr-1}*(1+{INFL})"
            g_guard = (f"IF(F{rr}<=0,0,"
                       f"IF({g_fixed}/F{rr}>{SWR}*(1+{BAND}),{g_fixed}*(1-{ADJ}),"
                       f"IF({g_fixed}/F{rr}<{SWR}*(1-{BAND}),{g_fixed}*(1+{ADJ}),{g_fixed})))")
            put(wsDd, f"G{rr}",
                f"=IF(A{rr}<{DRAW_YR},0,IF(A{rr}={DRAW_YR},{SWR}*F{rr},"
                f'IF({METHOD}="Fixed",{g_fixed},{g_guard})))', F_FX, MONEY)
        put(wsDd, f"B{rr}", f"=A{rr}-{BIRTH_R}", F_FX, YEAR)
        put(wsDd, f"F{rr}", f"=C{rr}+D{rr}+E{rr}", F_FX, MONEY)
        put(wsDd, f"H{rr}", f"=IF(B{rr}>={RMDAGE},D{rr}/INDEX({ULT_B},MATCH(B{rr},{ULT_A},0)),0)", F_FX, MONEY)
        put(wsDd, f"I{rr}", f"=MIN(H{rr},D{rr})", F_FX, MONEY)
        put(wsDd, f"J{rr}", f"=MIN(MAX(G{rr}-I{rr},0),C{rr})", F_FX, MONEY)
        put(wsDd, f"K{rr}", f"=MIN(MAX(G{rr}-I{rr}-J{rr},0),D{rr}-I{rr})", F_FX, MONEY)
        put(wsDd, f"L{rr}", f"=MIN(MAX(G{rr}-I{rr}-J{rr}-K{rr},0),E{rr})", F_FX, MONEY)
        put(wsDd, f"M{rr}", f"=MIN(G{rr},I{rr}+J{rr}+K{rr}+L{rr})", F_FX, MONEY)
        put(wsDd, f"N{rr}", f"=MAX(I{rr}-G{rr},0)", F_FX, MONEY)
        put(wsDd, f"O{rr}", f"=C{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsDd, f"P{rr}", f"=D{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsDd, f"Q{rr}", f"=E{rr}*{RET_RATE}", F_FX, MONEY)
        put(wsDd, f"R{rr}", f"=C{rr}+O{rr}-J{rr}+N{rr}", F_FX, MONEY)
        put(wsDd, f"S{rr}", f"=D{rr}+P{rr}-I{rr}-K{rr}", F_FX, MONEY)
        put(wsDd, f"T{rr}", f"=E{rr}+Q{rr}-L{rr}", F_FX, MONEY)
        put(wsDd, f"U{rr}", f"=R{rr}+S{rr}+T{rr}", F_FX, MONEY)
        put(wsDd, f"V{rr}", f"=U{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)
        put(wsDd, f"W{rr}", f"=M{rr}/(1+{INFL})^(A{rr}-{START})", F_FX, MONEY)
        put(wsDd, f"X{rr}", f"=J{rr}*{GAINF}", F_FX, MONEY)
        # Base year is the deduction record's own tax year, not a literal: a
        # superseding TY record re-bases this automatically.
        put(wsDd, f"Y{rr}", f"={WADED}*(1+{WAIDX})^(A{rr}-{WADED_TY})", F_FX, MONEY)
        put(wsDd, f"Z{rr}", f"=MAX(0,(X{rr}-Y{rr})*{WARATE})", F_FX, MONEY)
        put(wsDd, f"AA{rr}", f'=IF(Z{rr}>0,"CHECK","-")', F_FX)

    wsDd.freeze_panes = "C5"
    wsDd.column_dimensions["A"].width = 7
    wsDd.column_dimensions["B"].width = 5
    for i in range(3, 28):
        wsDd.column_dimensions[get_column_letter(i)].width = 13

    # =====================================================================
    # CHECKS
    # =====================================================================
    put(wsChk, "A1", "Checks - cross-foot every year, continuity, defect regressions, assumption register", F_TITLE)
    put(wsChk, "A2", "Cross-foot: End Total must equal Beg Total + growth + inflows - outflows for every year of every phase and scenario. PASS requires |delta| < $0.01.", F_NOTE)

    put(wsChk, "A4", "Year", F_HDR, None, FILL_HDR)
    put(wsChk, "B4", "Phase / scenario", F_HDR, None, FILL_HDR)
    put(wsChk, "C4", "Delta ($)", F_HDR, None, FILL_HDR)
    put(wsChk, "D4", "Status", F_HDR, None, FILL_HDR)

    chk_row = 5
    CHK_DATA_FIRST = chk_row
    for rr in range(ACC_FIRST, ACC_LAST + 1):
        put(wsChk, f"A{chk_row}", f"=Accumulation!A{rr}", F_LN, YEAR)
        put(wsChk, f"B{chk_row}", "Accumulation", F_FX)
        put(wsChk, f"C{chk_row}", f"=Accumulation!T{rr}-(Accumulation!F{rr}+Accumulation!N{rr}+Accumulation!O{rr}+Accumulation!P{rr}+Accumulation!L{rr})", F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1
    for rr in range(GA_FIRST, GA_LAST + 1):
        put(wsChk, f"A{chk_row}", f"=Gap!A{rr}", F_LN, YEAR)
        put(wsChk, f"B{chk_row}", "Gap - A growth only", F_FX)
        put(wsChk, f"C{chk_row}", f"=Gap!M{rr}-(Gap!F{rr}+Gap!G{rr}+Gap!H{rr}+Gap!I{rr})", F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1
    for rr in range(GB_FIRST, GB_LAST + 1):
        put(wsChk, f"A{chk_row}", f"=Gap!A{rr}", F_LN, YEAR)
        put(wsChk, f"B{chk_row}", "Gap - B early draws", F_FX)
        put(wsChk, f"C{chk_row}", f"=Gap!R{rr}-(Gap!F{rr}+Gap!L{rr}+Gap!M{rr}+Gap!N{rr}-Gap!K{rr})", F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1
    for rr in range(GC_FIRST, GC_LAST + 1):
        put(wsChk, f"A{chk_row}", f"=Gap!A{rr}", F_LN, YEAR)
        put(wsChk, f"B{chk_row}", "Gap - C conversions", F_FX)
        put(wsChk, f"C{chk_row}", f"=Gap!O{rr}-(Gap!F{rr}+Gap!I{rr}+Gap!J{rr}+Gap!K{rr}-Gap!H{rr})", F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1
    for rr in range(DD_FIRST, DD_LAST + 1):
        put(wsChk, f"A{chk_row}", f"=Drawdown!A{rr}", F_LN, YEAR)
        put(wsChk, f"B{chk_row}", "Drawdown", F_FX)
        put(wsChk, f"C{chk_row}", f"=Drawdown!U{rr}-(Drawdown!F{rr}+Drawdown!O{rr}+Drawdown!P{rr}+Drawdown!Q{rr}-Drawdown!M{rr})", F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1

    cont = [
        (f"Continuity: Gap {GAP_FIRST} beg = Accumulation {LAST_CONTRIB} end",
         f"=Gap!F{GA_FIRST}-Accumulation!T{ACC_LAST}"),
        (f"Continuity: Drawdown {DRAW_Y} beg = Gap handoff (selected scenario)",
         f"=Drawdown!F{DD_FIRST}-(Gap!B{FEED}+Gap!C{FEED}+Gap!D{FEED})"),
        (f"Continuity: Drawdown {END_Y} age = {END_AGE}",
         f"=Drawdown!B{DD_LAST}-{END_AGE}"),
    ]
    for name, f in cont:
        put(wsChk, f"A{chk_row}", "-", F_FX)
        put(wsChk, f"B{chk_row}", name, F_FX)
        put(wsChk, f"C{chk_row}", f, F_FX, MONEY2)
        put(wsChk, f"D{chk_row}", f'=IF(ABS(C{chk_row})<0.01,"PASS","FAIL")', F_FX)
        chk_row += 1
    put(wsChk, f"A{chk_row}", "-", F_FX)
    put(wsChk, f"B{chk_row}", f"Defect 1 regression: contribution escalates ({START_Y+1} > {START_Y})", F_FX)
    put(wsChk, f"C{chk_row}", f"=Accumulation!L{ACC_FIRST+1}-Accumulation!L{ACC_FIRST}", F_FX, MONEY2)
    put(wsChk, f"D{chk_row}", f'=IF(C{chk_row}>0,"PASS","FAIL")', F_FX)
    chk_row += 1
    put(wsChk, f"A{chk_row}", "-", F_FX)
    put(wsChk, f"B{chk_row}", f"Defect 2 regression: draws escalate under Fixed ({DRAW_Y+1} > {DRAW_Y})", F_FX)
    put(wsChk, f"C{chk_row}", f"=Drawdown!G{DD_FIRST+1}-Drawdown!G{DD_FIRST}", F_FX, MONEY2)
    put(wsChk, f"D{chk_row}", f'=IF({METHOD}<>"Fixed","n/a",IF(C{chk_row}>0,"PASS","FAIL"))', F_FX)
    chk_row += 1
    CHK_DATA_LAST = chk_row - 1

    put(wsChk, f"A{chk_row + 1}", "FAIL count:", F_HDR)
    put(wsChk, f"C{chk_row + 1}", f'=COUNTIF(D{CHK_DATA_FIRST}:D{CHK_DATA_LAST},"FAIL")', F_FX, YEAR)
    put(wsChk, f"D{chk_row + 1}", f'=IF(C{chk_row + 1}=0,"ALL CHECKS PASS","REVIEW FAILURES")', F_HDR, None, FILL_KEY)
    reg_start = chk_row + 4

    put(wsChk, f"A{reg_start - 1}", "ASSUMPTION REGISTER (every input, its tax-year label, and status)", F_SEC)
    put(wsChk, f"A{reg_start}", "Assumption", F_HDR, None, FILL_HDR)
    put(wsChk, f"B{reg_start}", "Value", F_HDR, None, FILL_HDR)
    put(wsChk, f"C{reg_start}", "Tax year / basis", F_HDR, None, FILL_HDR)
    put(wsChk, f"D{reg_start}", "Status", F_HDR, None, FILL_HDR)
    put(wsChk, f"E{reg_start}", "Source / note", F_HDR, None, FILL_HDR)

    WA_RATE_FROM_Y = dt.date.fromisoformat(
        reg.record_for(ID_WARATE, AS_OF)["valid_from"]).year

    register = [
        ("Nominal annual return", "=Assumptions!B15", "n/a", "User input", f"Original model ({pct(RET_V)}). Planning signal, not a forecast.", PCT2),
        ("Inflation", "=Assumptions!B16", "n/a", "Assumption", f"Matches original's intended (never-applied) {pct(ESC_V)} escalator.", PCT2),
        ("Birth year", "=Assumptions!B7", "n/a", "Derived", f"Age {START_Y - BIRTH} in {START_Y} per original model - confirm.", YEAR),
        ("Retirement age / last contribution year", "=Assumptions!B10", "n/a", "User input", f"Original contributions end {LAST_CONTRIB}.", YEAR),
        ("Draw start year", "=Assumptions!B11", "n/a", "User input", f"Original draws begin {DRAW_Y}.", YEAR),
        (f"Employee deferral {START_Y}", "=Assumptions!B20", f"TY{START_Y}", "Assumed split", f"Original gave only combined {usd0(EMP_V + ER_V)} - split is assumed.", MONEY),
        (f"Employer contribution {START_Y}", "=Assumptions!B21", f"TY{START_Y}", "Assumed split", "See above.", MONEY),
        ("Contribution escalation", "=Assumptions!B23", "n/a", "Assumption", "The repaired escalator (defect 1).", PCT2),
        (f"{sec_of(ID_402G)} deferral limit", "=Assumptions!B27", f"TY{tyear(ID_402G)}", conf(ID_402G), f"{pro(ID_402G, 'short_cite')}. {pro(ID_402G, 'code_cite')}.", MONEY),
        (f"Catch-up {AGE50}+", "=Assumptions!B28", f"TY{tyear(ID_CU50)}", conf(ID_CU50), f"{pro(ID_CU50, 'short_cite')}. {pro(ID_CU50, 'code_cite')}.", MONEY),
        (f"Enhanced catch-up {AGE60}-{AGE63}", "=Assumptions!B29", f"TY{tyear(ID_CU60)}", conf(ID_CU60), f"{pro(ID_CU60, 'short_cite')}.", MONEY),
        (f"{sec_of(ID_415C)} annual additions", "=Assumptions!B30", f"TY{tyear(ID_415C)}", conf(ID_415C), f"{pro(ID_415C, 'short_cite')}.", MONEY),
        ("Limit indexation", "=Assumptions!B31", "n/a", "Assumption", "Real limits move in $500-type steps; modeled smooth.", PCT2),
        ("Opening balances (3 buckets)", "=Assumptions!B38", f"{START_Y}", "From original", f"D5 = {num(OB_TAXDEF)} + {num(OB_TAXABLE)} + {num(OB_ROTHV)}; bucket mapping assumed.", MONEY),
        ("SWR method", "=Assumptions!B41", "n/a", "User choice", "Fixed (Bengen) vs Guardrails (Guyton-Klinger-style, simplified).", None),
        ("Initial withdrawal rate", "=Assumptions!B42", "n/a", "User input", "Bengen 4%; Morningstar 2025 base 3.9% (30-yr horizon).", PCT2),
        ("Legacy flat draw (not used)", "=Assumptions!B45", "n/a", "Reference", f"Original {usd_k(LEGACY_V)} nominal draw kept for comparison only (defects 2-3).", MONEY),
        ("Gap scenario selector", "=Assumptions!B48", "n/a", "User choice", "A growth-only / B early draws / C conversions.", None),
        ("Roth conversion per gap year", "=Assumptions!B49", "n/a", "User input", "Size with CPA against current bracket tables - not computed here.", MONEY),
        ("Effective tax rate on conversions", "=Assumptions!B50", "n/a", "Assumption", "Paid from taxable bucket.", PCT),
        ("Other MAGI income, gap years", "=Assumptions!B51", "n/a", "Assumption", "For the IRMAA proxy only.", MONEY),
        ("Filing status", "=Assumptions!B52", "n/a", "Confirm", "Skill default MFJ - confirm before results depend on it.", None),
        ("IRMAA tier-1 threshold", "=Assumptions!B53", f"{tyear(ID_IRMAA_MFJ)}", conf(ID_IRMAA_MFJ), f"{pro(ID_IRMAA_MFJ, 'short_cite')}. Per person, {VLOOKBACK:.0f}-yr lookback, cliff.", MONEY),
        ("IRMAA indexation", "=Assumptions!B54", "n/a", "Assumption", f"First four tiers indexed; fifth frozen to {val(ID_FREEZE):.0f}.", PCT2),
        ("WA excise rate", "=Assumptions!B57", f"TY{WA_RATE_FROM_Y}+", conf(ID_WARATE), f"{pro(ID_WARATE, 'short_cite')}; {pct(VWARATE + VWASUR)} surcharge above {usd_m(VWATHR)} gains NOT modeled.", PCT),
        ("WA CG standard deduction", "=Assumptions!B58", f"TY{WADED_TY}", conf(ID_WADED), f"TY{WADED_EXPECT_TY} amount must be confirmed at {pro(ID_WADED, 'issuer')}.", MONEY),
        ("Gain fraction of taxable withdrawals", "=Assumptions!B60", "n/a", "Assumption", "Share of taxable draws that is realized LTCG.", PCT),
        ("RMD start age", "=Assumptions!B64", "n/a", conf(ID_RMD), f"{pro(ID_RMD, 'short_cite')}: {RMD_AGE_HI:.0f} for born {RMD_COHORT:.0f}+. {pro(ID_RMD, 'code_cite')}.", YEAR),
        ("Uniform Lifetime Table divisors", f"=Assumptions!B{ULT_FIRST}", "n/a", conf(ULT_ID0), f"{pro(ULT_ID0, 'code_cite')} - verify vs IRS Pub. 590-B.", NUM1),
        ("Growth convention", "n/a", "n/a", "Structural", "Growth on beginning balances; flows post at year-end (matches original).", None),
        ("Draws are gross of income tax", "n/a", "n/a", "Structural", "Income tax on withdrawals not modeled - advisor/CPA item.", None),
        ("Social Security not modeled", "n/a", "n/a", "Structural", "Conservative omission; claiming strategy is an advisor item.", None),
        ("Single blended return across buckets", "n/a", "n/a", "Structural", "No asset-location differentiation - advisor item.", None),
    ]
    rr = reg_start + 1
    for name, v, ty, status, src, fmt in register:
        put(wsChk, f"A{rr}", name, F_FX)
        if isinstance(v, str) and v.startswith("="):
            put(wsChk, f"B{rr}", v, F_LN, fmt)
        else:
            put(wsChk, f"B{rr}", v, F_FX, fmt)
        put(wsChk, f"C{rr}", ty, F_FX)
        put(wsChk, f"D{rr}", status, F_FX)
        put(wsChk, f"E{rr}", src, F_NOTE, align=WRAP)
        rr += 1

    wsChk.column_dimensions["A"].width = 42
    wsChk.column_dimensions["B"].width = 34
    wsChk.column_dimensions["C"].width = 16
    wsChk.column_dimensions["D"].width = 20
    wsChk.column_dimensions["E"].width = 70

    # =====================================================================
    # ADVISOR
    # =====================================================================
    put(wsAdv, "A1", f"Take to your advisor - Retirement Strategy v2 (prepared {PREPARED})", F_TITLE)
    put(wsAdv, "A2", "This workbook is educational scenario analysis, not investment, tax, or legal advice. Each decision below is SIZED here and DECIDED with your fiduciary advisor / CPA / estate attorney.", F_NOTE, align=WRAP)
    put(wsAdv, "A3", "Citations are from the skill's authority index - text not re-verified this session.", F_NOTE)

    RMD_AGE_FOR_BIRTH = RMD_AGE_HI if BIRTH >= RMD_COHORT else RMD_AGE_LO
    adv_rows = [
        ("SEC", "LIVE FIGURES FROM THIS MODEL", None, None),
        ("KV", f"Projected balance at draw start ({DRAW_Y}, nominal)", f"=Drawdown!F{DD_FIRST}", MONEY),
        ("KV", f"First modeled draw ({DRAW_Y}, nominal)", f"=Drawdown!G{DD_FIRST}", MONEY),
        ("KV", f"First draw in {START_Y} purchasing power", f"=Drawdown!W{DD_FIRST}", MONEY),
        ("KV", f"Legacy {usd_k(LEGACY_V)} flat draw as % of {DRAW_Y} balance", f"=Assumptions!B45/Drawdown!F{DD_FIRST}", PCT2),
        ("TXT", "The legacy draw runs roughly double the research safe-withdrawal range (Bengen 4%, Morningstar 2025 base 3.9%) - the single most important conversation to have.", None, None),
        ("SEC", "DECISIONS PENDING (authority - fact - question)", None, None),
        ("TXT", f"1. ROTH CONVERSION WINDOW {GAP_FIRST}-{GAP_LAST} (Gap tab, Scenario C). Inputs: conversion/yr and effective rate from Assumptions. Ask: 'Model the proposed conversions (IRC 408A) against the IRMAA tiers under {pro(ID_IRMAA_MFJ, 'code_cite')} - {word(VLOOKBACK)}-year lookback, per person, cliff pricing - before we size any year; confirm bracket headroom against current-year tables.' Note the widow's-penalty and RMD-bulge motivations.", None, None),
        ("TXT", f"2. WITHDRAWAL METHOD (Drawdown tab). {SWR_METHOD} {pct(SWR_V)} inflation-adjusted vs guardrails (band +/-{pct(BAND_V)}, adjust {pct(ADJ_V)}). Ask: 'Given the projected surplus/failure profile, is a dynamic rule better than static - and what spending floor do we need?'", None, None),
        ("TXT", f"3. RMD COORDINATION (born {BIRTH} - RMD age {RMD_AGE_FOR_BIRTH:.0f}, first RMD {BIRTH + int(RMD_AGE_FOR_BIRTH)}; {pro(ID_RMD, 'short_cite')} / {pro(ID_RMD, 'code_cite')}). First-RMD April-1 deferral stacks two RMDs into one tax year (bracket + IRMAA risk). QCDs available from age {num(VQCDAGE)} (up to {usd0(VQCD)} TY{tyear(ID_QCD)}, indexed; {pro(ID_QCD, 'code_cite')}) - an AGI exclusion that beats a deduction for IRMAA purposes.", None, None),
        ("TXT", f"4. WASHINGTON OVERLAY. Capital-gains excise ({pro(ID_WARATE, 'short_cite')}): {pct(VWARATE)} on WA LTCG above the standard deduction ({usd0(VWADED)} TY{WADED_TY} - {conf(ID_WADED)} TY{WADED_EXPECT_TY} at DOR); retirement accounts exempt - raises asset-location stakes. Estate: WA exclusion ~{usd_m(VWAEST)} ({pro(ID_WAEST, 'code_cite')}), NO portability - credit-shelter/disclaimer design is an estate-attorney item. WA Cares (RCW 50B.04) is a partial LTC offset only.", None, None),
        ("TXT", f"5. CONTRIBUTION STRUCTURE. Employee/employer split here is ASSUMED ({num(EMP_V)} / {num(ER_V)} = original {num(EMP_V + ER_V)}) - confirm actuals. Limits: {pro(ID_402G, 'code_cite')} {usd0(V402G)} TY{tyear(ID_402G)} shared across employers; {sec_of(ID_415C)} {usd0(V415C)} per unrelated employer; catch-up {AGE50}+ {usd0(VCU50)}; enhanced {AGE60}-{AGE63} {usd0(VCU60)} (years {BIRTH + int(AGE60)}-{BIRTH + int(AGE63)} for you). From {tyear(ID_ROTHCU)}, catch-up must be Roth if prior-year FICA wages from that employer exceeded {usd0(VROTHCU)} ({pro(ID_ROTHCU, 'short_cite')}).", None, None),
        ("TXT", f"6. SOCIAL SECURITY - NOT MODELED (conservative). Claim window {VCLAIM_LO:.0f}-{VCLAIM_HI:.0f}, FRA {FRA_AGE:.0f} (born {FRA_COHORT:.0f}+); delaying is longevity insurance (about 8%/yr FRA to {VCLAIM_HI:.0f}). Ask how benefits integrate with the draw schedule and benefit taxation (IRC 86).", None, None),
        ("SEC", "AUTHORITIES IN PLAY (index citations - not re-verified this session)", None, None),
        ("TXT", f"IRC 402(g); 414(v); 415(c); 408A; 401(a)(9) and Treas. Reg. 1.401(a)(9)-9; 408(d)(8); 86; 1411 (NIIT - conversions raise MAGI over the {usd_k(VNIIT)} MFJ gate for other investment income); SECURE 2.0 secs. 107, 109, 126, 603; 42 U.S.C. 1395r(i) (IRMAA); RCW 82.87; RCW 83.100; RCW 26.16 (community property - double step-up under IRC 1014(b)(6)); RCW 50B.04.", None, None),
        ("SEC", "BRING TO THE MEETING", None, None),
        ("TXT", "This workbook (Assumptions + Checks tabs first); latest 401(k)/IRA/Roth statements confirming the three bucket balances; actual employee/employer contribution split; prior-year tax return (for conversion headroom and IRMAA base); advisor's Form ADV Part 2 / CRS if evaluating a new advisor.", None, None),
        ("SEC", "WHAT CHANGED VS THE ORIGINAL MODEL", None, None),
        ("TXT", f"(1) Escalator repaired - contributions now grow at a stated rate from a named input instead of compounding a zero cell. (2) Draws now inflation-escalate from a withdrawal RATE instead of a flat nominal {usd_k(LEGACY_V)}. (3) 'Withdrawal Rate' label no longer holds a dollar amount - rate and legacy dollar figure are separate labeled cells. (4) The {GAP_FIRST}-{GAP_LAST} gap is modeled explicitly under three scenarios instead of silently compounding.", None, None),
    ]
    rr = 5
    for kind, text, formula, fmt in adv_rows:
        if kind == "SEC":
            put(wsAdv, f"A{rr}", text, F_SEC)
        elif kind == "KV":
            put(wsAdv, f"A{rr}", text, F_FX)
            put(wsAdv, f"B{rr}", formula, F_LN, fmt)
        else:
            put(wsAdv, f"A{rr}", text, F_FX, align=WRAP)
            wsAdv.row_dimensions[rr].height = 45
        rr += 1
    wsAdv.column_dimensions["A"].width = 110
    wsAdv.column_dimensions["B"].width = 18

    wb.save(out_path)
    return {
        "out": str(out_path),
        "layout": (f"ACC {ACC_FIRST}-{ACC_LAST}; GA {GA_FIRST}-{GA_LAST}; "
                   f"GB {GB_FIRST}-{GB_LAST}; GC {GC_FIRST}-{GC_LAST}; "
                   f"DD {DD_FIRST}-{DD_LAST}; CHK data {CHK_DATA_FIRST}-{CHK_DATA_LAST}"),
    }


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--inputs", default=None,
                    help="inputs JSON; defaults to inputs.local.json, "
                         "falling back to inputs.example.json")
    ap.add_argument("--parameters", default=None)
    a = ap.parse_args(argv)
    inputs = a.inputs
    if inputs is None:
        inputs = DEFAULT_INPUTS if DEFAULT_INPUTS.exists() else EXAMPLE_INPUTS
    result = build(a.out, inputs, a.parameters)
    print("saved", result["out"])
    print("layout:", result["layout"])
    return 0


if __name__ == "__main__":
    sys.exit(main())
