"""The gate: prove build.py reproduces the reference engine cell for cell.

Compares CONTENT, not container bytes. A naive byte diff of two .xlsx files
fails dishonestly -- the zip container records timestamps, so two identical
workbooks built a second apart differ. So this walks the workbooks:

  primary    every cell in the union of both used ranges: .value (a formula is
             its string) and .number_format
  structural sheet names AND order, column widths, row heights, freeze panes,
             data-validation ranges and lists

Both passes must be empty for the gate to pass.

Exit codes are deliberately three-valued:
  0  identical -- the gate passes
  1  differences found, or a build failed
  2  SKIPPED: the reference fixture is absent

Two is not zero on purpose. build_v2.py is gitignored, so a fresh clone does
not have it; if a missing fixture exited 0 this test would report success
while comparing nothing.

Privacy: failure output names a sheet, an address, and a category. Household
figures -- balances, birth year, the contribution split, the gap-year amounts,
and sums derived from them -- are withheld, in numeric cells and in prose that
embeds them alike. An identity test that dumps balances into a terminal or a CI
log defeats the reason the inputs were split out in the first place. Legal
parameters ARE printed: they are public, and hiding them would conceal exactly
what a reviewer needs.

Known residual: derived YEARS stay visible ("Accumulation 2026-2052"). They are
arithmetic on the birth year, so a determined reader could work backwards.
Withholding them too would blank almost every title and label and make failures
unreadable; the figures themselves never appear.
"""
from __future__ import annotations

import subprocess
import sys
import tempfile
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
REFERENCE = HERE / "build_v2.py"
CANDIDATE = HERE / "build.py"
EXAMPLE_INPUTS = HERE / "inputs.example.json"

EXIT_PASS, EXIT_FAIL, EXIT_SKIP = 0, 1, 2

# Which input sections carry values that must never reach a terminal or a CI
# log. Balances, birth year, the contribution split, and the gap-year figures
# describe a household. Timeline years and assumption rates do not, and
# withholding them would make failures unreadable for no privacy gain.
SENSITIVE_SECTIONS = ("person", "balances", "contributions", "gap")
SENSITIVE_EXTRA = (("swr", "legacy_flat_draw"),)


def plane_value_sets():
    """Numbers that are safe to print, and numbers that are not.

    Legal parameters are public -- withholding them would hide exactly what a
    reviewer needs. Household figures are not. The test cannot tell them apart
    by cell address without duplicating the engine's layout, so it asks the
    data. Ties resolve to withholding.
    """
    import json
    legal, sensitive = set(), set()

    def walk(node, sink):
        if isinstance(node, dict):
            for v in node.values():
                walk(v, sink)
        elif isinstance(node, list):
            for v in node:
                walk(v, sink)
        elif isinstance(node, (int, float)) and not isinstance(node, bool):
            sink.add(float(node))

    params = HERE / "parameters.json"
    if params.exists():
        for rec in json.loads(params.read_text(encoding="ascii"))["records"]:
            walk(rec["payload"], legal)
    # COUPLING: this must read the SAME inputs file the comparison builds from
    # (EXAMPLE_INPUTS, below in main()). Derive the withhold-list from one file
    # and build from another and the redaction silently stops covering the
    # values actually in the workbook.
    if EXAMPLE_INPUTS.exists():
        inp = json.loads(EXAMPLE_INPUTS.read_text(encoding="utf-8"))
        for section in SENSITIVE_SECTIONS:
            got = set()
            walk(inp.get(section, {}), got)
            sensitive |= got
            # The workbook also prints SUMS of these -- the combined
            # contribution and the total of the three buckets. A derived figure
            # is as personal as the parts it came from.
            if len(got) > 1:
                sensitive.add(sum(got))
        for section, key in SENSITIVE_EXTRA:
            walk(inp.get(section, {}).get(key), sensitive)
    return legal, sensitive


def rendered_forms(values):
    """Every shape a sensitive number takes in the workbook's prose, so a
    string carrying one can be recognised and withheld."""
    out = set()
    for v in values:
        out.add(f"{v:,.0f}")
        two = f"{v:,.2f}"
        out.add(two.rstrip("0").rstrip(".") if "." in two else two)
        if float(v).is_integer():
            out.add(str(int(v)))
    return {s for s in out if len(s) >= 3}


LEGAL_VALUES, SENSITIVE_VALUES = plane_value_sets()
SENSITIVE_STRINGS = rendered_forms(SENSITIVE_VALUES)


def reference_output_path() -> Path:
    """The reference engine writes to a path hard-coded in its own source. Read
    it rather than assume it -- and rather than edit a fixture that must stay
    untouched."""
    for line in REFERENCE.read_text(encoding="utf-8").splitlines():
        if line.startswith("OUT"):
            _, _, rhs = line.partition("=")
            return Path(rhs.strip().strip('r"').strip('"').strip("'"))
    raise SystemExit("could not find OUT in the reference engine")


def run(cmd) -> None:
    proc = subprocess.run(cmd, capture_output=True, text=True)
    if proc.returncode != 0:
        print(f"build failed: {' '.join(str(c) for c in cmd)}")
        print(proc.stdout[-2000:])
        print(proc.stderr[-2000:])
        raise SystemExit(EXIT_FAIL)


def used_cells(ws):
    seen = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None or c.number_format != "General":
                seen[c.coordinate] = c
    return seen


def compare_cells(ref_wb, new_wb):
    diffs, compared = [], 0
    for name in ref_wb.sheetnames:
        ref_ws, new_ws = ref_wb[name], new_wb[name]
        ref_cells, new_cells = used_cells(ref_ws), used_cells(new_ws)
        for addr in sorted(set(ref_cells) | set(new_cells)):
            compared += 1
            a = ref_cells.get(addr)
            b = new_cells.get(addr)
            if a is None or b is None:
                diffs.append((name, addr, "missing",
                              "present only in reference" if b is None
                              else "present only in candidate"))
                continue
            if a.value != b.value:
                diffs.append((name, addr, "value", redact(name, a.value, b.value)))
            if a.number_format != b.number_format:
                diffs.append((name, addr, "number_format",
                              f"{a.number_format!r} vs {b.number_format!r}"))
    return diffs, compared


def redact(sheet, ref_value, new_value):
    """Formulas and labels are safe to show and are what a failure needs.
    Numeric literals are shown only when they are demonstrably legal
    parameters."""
    if ref_value is None or isinstance(ref_value, bool):
        return f"{ref_value!r} vs {new_value!r}"
    if isinstance(ref_value, str):
        both = f"{ref_value} {new_value}"
        if any(s in both for s in SENSITIVE_STRINGS):
            return "string differs (contents withheld: embeds household figures)"
        return f"{ref_value!r} vs {new_value!r}"
    v = float(ref_value)
    if v in SENSITIVE_VALUES:
        return "numeric value differs (contents withheld: household figure)"
    if v in LEGAL_VALUES:
        return f"{ref_value!r} vs {new_value!r}"
    return "numeric value differs (contents withheld: not a legal parameter)"


def compare_structure(ref_wb, new_wb):
    diffs = []
    if ref_wb.sheetnames != new_wb.sheetnames:
        diffs.append(("<workbook>", "-", "structure",
                      f"sheet names/order {ref_wb.sheetnames} vs {new_wb.sheetnames}"))
        return diffs
    for name in ref_wb.sheetnames:
        a, b = ref_wb[name], new_wb[name]
        if a.freeze_panes != b.freeze_panes:
            diffs.append((name, "-", "structure",
                          f"freeze_panes {a.freeze_panes} vs {b.freeze_panes}"))
        for key in sorted(set(a.column_dimensions) | set(b.column_dimensions)):
            wa = a.column_dimensions[key].width if key in a.column_dimensions else None
            wb_ = b.column_dimensions[key].width if key in b.column_dimensions else None
            if wa != wb_:
                diffs.append((name, f"col {key}", "structure", f"width {wa} vs {wb_}"))
        for key in sorted(set(a.row_dimensions) | set(b.row_dimensions)):
            ha = a.row_dimensions[key].height if key in a.row_dimensions else None
            hb = b.row_dimensions[key].height if key in b.row_dimensions else None
            if ha != hb:
                diffs.append((name, f"row {key}", "structure", f"height {ha} vs {hb}"))
        av = sorted((dv.sqref.__str__(), dv.formula1)
                    for dv in a.data_validations.dataValidation)
        bv = sorted((dv.sqref.__str__(), dv.formula1)
                    for dv in b.data_validations.dataValidation)
        if av != bv:
            diffs.append((name, "-", "structure", f"data validations {av} vs {bv}"))
    return diffs


def main() -> int:
    if not REFERENCE.exists():
        print("SKIPPED (reference fixture absent)")
        print(f"  {REFERENCE} is gitignored and not present in this checkout.")
        print("  Exiting 2 -- a skip is not a pass.")
        return EXIT_SKIP
    if not EXAMPLE_INPUTS.exists():
        print(f"SKIPPED (missing {EXAMPLE_INPUTS.name})")
        return EXIT_SKIP

    ref_out = reference_output_path()
    with tempfile.TemporaryDirectory() as tmp:
        new_out = Path(tmp) / "candidate.xlsx"
        run([sys.executable, str(REFERENCE)])
        run([sys.executable, str(CANDIDATE),
             "--inputs", str(EXAMPLE_INPUTS), "--out", str(new_out)])

        ref_wb = openpyxl.load_workbook(ref_out, data_only=False)
        new_wb = openpyxl.load_workbook(new_out, data_only=False)

        struct = compare_structure(ref_wb, new_wb)
        cells, compared = ([], 0) if struct and struct[0][0] == "<workbook>" \
            else compare_cells(ref_wb, new_wb)

    print(f"sheets compared:  {len(ref_wb.sheetnames)} "
          f"({', '.join(ref_wb.sheetnames)})")
    print(f"cells compared:   {compared}")
    print(f"cell differences: {len(cells)}")
    print(f"structural diffs: {len(struct)}")

    if not cells and not struct:
        print("\nIDENTICAL -- gate PASSES at zero cell-level differences.")
        return EXIT_PASS

    print("\nDIFFERENCES")
    for sheet, addr, kind, detail in (cells + struct)[:200]:
        print(f"  {sheet}!{addr}  [{kind}]  {detail}")
    if len(cells) + len(struct) > 200:
        print(f"  ... {len(cells) + len(struct) - 200} more")
    return EXIT_FAIL


if __name__ == "__main__":
    sys.exit(main())
